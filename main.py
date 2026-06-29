"""
AI 学术文献相关性评分工具 - 核心模块
使用 OpenAI 嵌入模型计算文献与研究主题的语义相似度
"""

import pandas as pd
import numpy as np
import os
import time
import pickle
import json
import httpx
from tqdm import tqdm
from openai import OpenAI, APIConnectionError, AuthenticationError, APIError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def normalize_api_key(api_key):
    for line in str(api_key).splitlines():
        line = line.strip()
        if line:
            if line.startswith("Bearer "):
                line = line[len("Bearer "):]
            return line.strip()
    return ""


def make_openai_client(api_key, base_url, timeout=60):
    return OpenAI(
        api_key=normalize_api_key(api_key),
        base_url=base_url,
        timeout=timeout,
        http_client=httpx.Client(trust_env=False),
    )


def ai_score_ref_column(
    input_file,
    output_file,
    api_key,
    topic,
    base_url="https://aihubmix.com/v1",
    model="text-embedding-3-small",
    batch_size=20,
    delay=1.5,
    cache_file="embedding_cache.pkl",
    llm_screening=True,
    llm_model="gpt-4o-mini",
    llm_score_threshold=0.45,
    llm_delay=0.2,
    progress_callback=None,
    cancel_check=None,
):
    """
    AI 评分核心函数 - 为参考文献计算与主题的相关性分数

    Args:
        input_file: 输入 Excel 文件路径（必须包含 'ref' 列）
        output_file: 输出 Excel 文件路径
        api_key: OpenAI API 密钥
        topic: 研究主题文本
        base_url: API 基础 URL
        model: 嵌入模型名称
        batch_size: 批处理大小
        delay: 请求间隔（秒）
        cache_file: 嵌入缓存文件路径
        llm_screening: 是否对低相似度文献启用 LLM scope 筛选
        llm_model: 用于 scope 筛选的聊天模型
        llm_score_threshold: 低于该相似度时调用 LLM 做二次判断
        llm_delay: LLM 请求间隔（秒）
        progress_callback: 进度回调函数 callback(progress_pct, status_message, done=False)
        cancel_check: 取消检查函数，返回 True 表示应停止
    """
    client = make_openai_client(
        api_key=api_key,
        base_url=base_url,
        timeout=60
    )

    df = pd.read_excel(input_file)

    # ===== 识别 ref / refs / Reference 列（不区分大小写）=====
    ref_aliases = {"ref", "refs", "reference"}
    found_col = None
    for col in df.columns:
        if col.lower().strip() in ref_aliases:
            found_col = col
            break
    if found_col is None:
        raise ValueError("Excel 必须包含 'ref'、'refs' 或 'Reference' 列（不区分大小写）")
    if found_col != "ref":
        df.rename(columns={found_col: "ref"}, inplace=True)

    # ===== 文本标准化（提高去重率）=====
    def normalize(t):
        return str(t).strip().lower()

    texts_raw = df["ref"].astype(str).tolist()
    texts_norm = [normalize(t) for t in texts_raw]

    # ===== 加载缓存 =====
    if os.path.exists(cache_file):
        with open(cache_file, "rb") as f:
            cache = pickle.load(f)
    else:
        cache = {}

    # ===== 相似度 =====
    def cosine_similarity(a, b):
        return np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b))

    # ===== embedding（带缓存 + 重试）=====
    def get_embeddings_batch(texts, max_retries=5):
        uncached = [t for t in texts if t not in cache]

        if uncached:
            for attempt in range(max_retries):
                try:
                    response = client.embeddings.create(
                        model=model,
                        input=uncached
                    )

                    for t, emb in zip(uncached, response.data):
                        cache[t] = np.array(emb.embedding)

                    break

                except Exception as e:
                    wait = 2 ** attempt
                    msg = f"[Retry {attempt+1}] waiting {wait}s..."
                    print(msg)
                    if progress_callback:
                        progress_callback(0, msg)
                    time.sleep(wait)
            else:
                raise Exception("Embedding failed")

        return [cache[t] for t in texts]

    # ===== LLM scope 筛选 =====
    def parse_json_object(text):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            start = text.find("{")
            end = text.rfind("}")
            if start >= 0 and end > start:
                return json.loads(text[start:end + 1])
            raise

    def llm_scope_check(reference, max_retries=3):
        system_prompt = (
            "You screen manuscript titles or references for a special issue. "
            "Be conservative: only mark out_of_scope when it is clearly outside the scope. "
            "If the title is plausibly related to any meaningful part of the scope, mark in_scope. "
            "If there is not enough information, mark uncertain."
        )
        user_prompt = f"""
Special issue title / scope:
{topic}

Manuscript title or reference:
{reference}

Task:
Determine whether the manuscript is within the academic scope of the special issue.

Rules:
- The manuscript does not need to match every keyword in the scope.
- It is enough if it belongs to one meaningful sub-area of the scope.
- Use out_of_scope only when the manuscript is clearly unrelated.
- Use uncertain when the relationship is weak, ambiguous, or the title lacks enough information.

Return only JSON with this schema:
{{
  "decision": "in_scope | out_of_scope | uncertain",
  "confidence": "high | medium | low",
  "reason": "short explanation"
}}
"""
        for attempt in range(max_retries):
            try:
                request = {
                    "model": llm_model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    "temperature": 0,
                }
                try:
                    response = client.chat.completions.create(
                        **request,
                        response_format={"type": "json_object"},
                    )
                except Exception as e:
                    if "response_format" not in str(e):
                        raise
                    response = client.chat.completions.create(**request)
                content = response.choices[0].message.content or "{}"
                data = parse_json_object(content)
                decision = str(data.get("decision", "uncertain")).strip().lower()
                confidence = str(data.get("confidence", "low")).strip().lower()
                reason = str(data.get("reason", "")).strip()

                if decision not in {"in_scope", "out_of_scope", "uncertain"}:
                    decision = "uncertain"
                if confidence not in {"high", "medium", "low"}:
                    confidence = "low"
                return decision, confidence, reason
            except Exception as e:
                wait = 2 ** attempt
                msg = f"[LLM Retry {attempt+1}] waiting {wait}s..."
                print(msg)
                if progress_callback:
                    progress_callback(0, msg)
                time.sleep(wait)

        return "uncertain", "low", "LLM screening failed; kept by conservative rule."

    # ===== 全局去重（核心）=====
    text_to_indices = {}
    for idx, t in enumerate(texts_norm):
        text_to_indices.setdefault(t, []).append(idx)

    unique_texts = list(text_to_indices.keys())
    total = len(unique_texts)

    print(f"原始数量：{len(texts_norm)}")
    print(f"去重后：{total}")

    if progress_callback:
        progress_callback(0, f"原始数量：{len(texts_norm)}, 去重后：{total}")

    # ===== topic embedding =====
    topic_emb = get_embeddings_batch([normalize(topic)])[0]

    # ===== 主循环 =====
    scores = [None] * len(df)
    scope_decisions = ["pending"] * len(df)
    scope_confidences = [""] * len(df)
    excludes = [False] * len(df)
    screening_reasons = [""] * len(df)
    processed = 0

    for i in range(0, total, batch_size):
        # 检查是否取消
        if cancel_check and cancel_check():
            print("用户取消操作")
            return None

        batch = unique_texts[i:i+batch_size]

        try:
            embs = get_embeddings_batch(batch)
        except:
            embs = [get_embeddings_batch([t])[0] for t in batch]

        for text, emb in zip(batch, embs):
            sim = cosine_similarity(topic_emb, emb)
            # 回填所有原始位置
            for idx in text_to_indices[text]:
                scores[idx] = sim

        processed += len(batch)
        progress = (processed / total) * (70 if llm_screening else 100)

        if progress_callback:
            progress_callback(progress, f"处理中：{processed}/{total} ({progress:.1f}%)")

        time.sleep(delay)

        # 断点保存
        if i % 200 == 0:
            df_temp = df.copy()
            df_temp["score"] = scores
            df_temp.to_excel("progress.xlsx", index=False)

    # ===== 低分文献 LLM scope 筛选 =====
    llm_candidates = [
        text for text in unique_texts
        if scores[text_to_indices[text][0]] is not None
        and scores[text_to_indices[text][0]] < llm_score_threshold
    ]

    if not llm_screening:
        llm_candidates = []

    checked = 0
    total_llm = len(llm_candidates)

    for text in unique_texts:
        idx0 = text_to_indices[text][0]
        score = scores[idx0]
        if score is not None and score >= llm_score_threshold:
            for idx in text_to_indices[text]:
                scope_decisions[idx] = "in_scope_by_score"
                scope_confidences[idx] = "not_checked"
                excludes[idx] = False
                screening_reasons[idx] = f"Embedding score >= {llm_score_threshold:.2f}; kept without LLM screening."

    if total_llm and progress_callback:
        progress_callback(70, f"开始 LLM scope 筛选：{total_llm} 条低分文献")

    for text in llm_candidates:
        if cancel_check and cancel_check():
            print("用户取消操作")
            return None

        decision, confidence, reason = llm_scope_check(texts_raw[text_to_indices[text][0]])
        exclude = decision == "out_of_scope" and confidence == "high"

        for idx in text_to_indices[text]:
            scope_decisions[idx] = decision
            scope_confidences[idx] = confidence
            excludes[idx] = exclude
            screening_reasons[idx] = reason

        checked += 1
        progress = 70 + (checked / total_llm) * 25
        if progress_callback:
            progress_callback(progress, f"LLM 筛选中：{checked}/{total_llm} ({progress:.1f}%)")

        time.sleep(llm_delay)

    for idx, decision in enumerate(scope_decisions):
        if decision == "pending":
            scope_decisions[idx] = "uncertain"
            scope_confidences[idx] = "low"
            excludes[idx] = False
            screening_reasons[idx] = "No screening decision was produced; kept by conservative rule."

    # ===== 写入结果（不排序）=====
    df["score"] = scores
    df["scope_decision"] = scope_decisions
    df["scope_confidence"] = scope_confidences
    df["exclude"] = excludes
    df["screening_reason"] = screening_reasons
    df.to_excel(output_file, index=False)

    # ===== 保存缓存 =====
    with open(cache_file, "wb") as f:
        pickle.dump(cache, f)

    if progress_callback:
        progress_callback(100, "完成！", done=True)

    print(f"完成！结果已保存到：{output_file}")


def validate_api_connection(api_key, base_url, model):
    """验证 API 连接"""
    try:
        client = make_openai_client(
            api_key=api_key,
            base_url=base_url,
            timeout=30
        )
        client.embeddings.create(
            model=model,
            input=["validate"]
        )
        print("[OK] API 连接验证通过")
        return True
    except AuthenticationError:
        print("[ERROR] API Key 无效或已过期")
        return False
    except APIConnectionError:
        print(f"[ERROR] 无法连接到 API 服务：{base_url}")
        print("  请检查：网络连接、API 服务状态、防火墙/代理设置")
        return False
    except Exception as e:
        print(f"[ERROR] API 验证失败：{str(e)}")
        return False


# ===== 国家区域高亮数据 =====
zone1 = {
    "Australia","Austria","Belgium","Bulgaria","Canada","Croatia","Cyprus",
    "Czech Republic","Denmark","Estonia","Finland","France","Germany","Greece",
    "Hungary","Iceland","Ireland","Israel","Italy","Japan","Latvia",
    "Liechtenstein","Lithuania","Luxembourg","Malta","Netherlands",
    "New Zealand","Norway","Poland","Portugal","Romania","Singapore",
    "Slovakia","Slovenia","South Korea","Republic of Korea","Spain","Sweden","Switzerland",
    "Taiwan","United Kingdom","UK","United States","USA"
}

zone2 = {
    "Argentina","Bahrain","Belarus","Brazil","Chile","Colombia","Cuba",
    "Ecuador","French Guiana","Guyana","Mexico","Morocco",
    "North Macedonia","Oman","Peru","Qatar","Serbia",
    "South Africa","Uruguay"
}

country_to_cctld = {
    "Australia": ".au","Austria": ".at","Belgium": ".be","Bulgaria": ".bg",
    "Canada": ".ca","Croatia": ".hr","Cyprus": ".cy","Czech Republic": ".cz",
    "Denmark": ".dk","Estonia": ".ee","Finland": ".fi","France": ".fr",
    "Germany": ".de","Greece": ".gr","Hungary": ".hu","Iceland": ".is",
    "Ireland": ".ie","Israel": ".il","Italy": ".it","Japan": ".jp",
    "Latvia": ".lv","Liechtenstein": ".li","Lithuania": ".lt","Luxembourg": ".lu",
    "Malta": ".mt","Netherlands": ".nl","New Zealand": ".nz","Norway": ".no",
    "Poland": ".pl","Portugal": ".pt","Romania": ".ro","Singapore": ".sg",
    "Slovakia": ".sk","Slovenia": ".si","South Korea": ".kr","Republic of Korea": ".kr",
    "Spain": ".es","Sweden": ".se","Switzerland": ".ch","Taiwan": ".tw",
    "United Kingdom": ".uk","UK": ".uk","United States": ".us","USA": ".us",
}

generic_domains = {".com", ".org", ".net", ".edu", ".gov", ".mil", ".io", ".co"}
regional_domains = {".eu", ".asia", ".int"}


def _clean(x):
    if x is None:
        return ""
    return str(x).replace("\xa0", "").strip()


def _get_email_domain(email):
    """从 email 提取国家域名后缀"""
    if not email or "@" not in email:
        return None
    domain = email.split("@")[-1].lower()
    if "." in domain:
        parts = domain.split(".")
        if len(parts) >= 2:
            if parts[-2] in ["co", "com", "ac", "gov", "mil", "edu"]:
                return "." + parts[-1]
            return "." + parts[-1]
    return None


def _check_country_email_match(country, email):
    """检查国家与 email 域名是否匹配"""
    if not email or not country:
        return "unknown"
    domain_suffix = _get_email_domain(email)
    if not domain_suffix:
        return "unknown"
    expected_cctld = country_to_cctld.get(country)
    if not expected_cctld:
        return "unknown"
    if domain_suffix == expected_cctld:
        return "match"
    if domain_suffix in generic_domains or domain_suffix in regional_domains:
        return "suspicious"
    return "mismatch"


def highlight_countries(file_path):
    """对 Excel 文件执行国家区域高亮（原地修改）"""
    wb = load_workbook(file_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    # 识别 Country/Region 列
    country_col_idx = None
    for i, h in enumerate(header):
        if h and h.strip().lower() in {"country/region", "country", "country region"}:
            country_col_idx = i
            break

    # 识别 email 列（兼容 EMail / email / E-mail）
    email_col_idx = None
    for i, h in enumerate(header):
        if h and h.strip().lower() in {"email", "e-mail", "mail"}:
            email_col_idx = i
            break

    if country_col_idx is None:
        print("[HighLight] 未找到 Country/Region 列，跳过高亮")
        wb.close()
        return

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue   = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        country_cell = row[country_col_idx]
        country = _clean(country_cell.value)

        if country in zone1:
            email = ""
            if email_col_idx is not None:
                email_cell = row[email_col_idx]
                email = _clean(email_cell.value)

            match_status = _check_country_email_match(country, email)
            if match_status in ["mismatch", "suspicious"]:
                country_cell.fill = orange
                if email_col_idx is not None:
                    row[email_col_idx].fill = orange
            else:
                country_cell.fill = yellow
        elif country in zone2:
            country_cell.fill = blue

    wb.save(file_path)
    wb.close()
    print(f"[HighLight] 国家区域高亮完成: {file_path}")


if __name__ == "__main__":
    # 运行前验证 API 连接
    # 从当前目录 key.txt 读取 API Key
    key_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "key.txt")
    api_key = ""
    if os.path.exists(key_file):
        with open(key_file, "r") as f:
            api_key = f.read().strip()
    if not api_key:
        print("[ERROR] 未找到 key.txt 或文件为空，请在当前目录创建 key.txt 并写入 API Key")
        exit(1)
    base_url = "https://aihubmix.com/v1"
    model = "text-embedding-3-small"

    print("正在验证 API 连接...")
    if not validate_api_connection(api_key, base_url, model):
        print("\nAPI 连接失败，请检查配置后重试")
        exit(1)

    ai_score_ref_column(
        input_file="input.xlsx",
        output_file="result.xlsx",
        api_key=api_key,
        topic="""
AI-enabled architectures for 5G and beyond networks;
Intelligent signal processing for wireless communications;
Model-aware and learning-enhanced wireless systems;
Intelligent resource management and network optimization;
AI-enabled applications, including joint communication and sensing, localization and other emerging 5G- and 6G-oriented scenarios.
"""
    )

    # 评分完成后自动执行国家区域高亮
    print("\n正在执行国家区域高亮标记...")
    highlight_countries("result.xlsx")
